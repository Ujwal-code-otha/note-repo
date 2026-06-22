import os
import sys
import json
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def main():
    if len(sys.argv) < 4:
        print("Usage: python style_excel_report.py <json_path> <excel_path> <title>")
        sys.exit(1)
        
    json_path = sys.argv[1]
    excel_path = sys.argv[2]
    title_text = sys.argv[3]
    
    if not os.path.exists(json_path):
        print(f"Error: JSON file not found at {json_path}")
        sys.exit(1)
        
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    records = data.get("testCases", [])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Execution Report"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # 1. Title Row
    ws.cell(row=1, column=1, value=f"{title_text} ({len(records)} Test Cases)")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=16, bold=True)
    
    # 2. Metadata Row
    gen_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row=2, column=1, value=f"Generated: {gen_time}")
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=11, italic=True)
    
    # 3. Empty Row 3
    
    # 4. Table Headers Row 4
    headers = ["#", "Test Case", "Category", "Status", "Duration (s)", "Details", "Timestamp"]
    ws.append([]) # Row 3
    ws.append(headers) # Row 4
    
    # Style header row
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    # Dark Green header fill like the image
    header_fill = PatternFill(start_color="134F2C", end_color="134F2C", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    for col_idx in range(1, 8):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if col_idx in [1, 4, 5, 7] else left_align
        
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Styles for status
    pass_font = Font(name="Calibri", size=11, bold=True, color="0F5132")
    pass_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    
    fail_font = Font(name="Calibri", size=11, bold=True, color="842029")
    fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    normal_font = Font(name="Calibri", size=11)
    
    # 5. Populate Data
    for idx, r in enumerate(records, start=1):
        row_num = 4 + idx
        
        status_val = r.get("status", r.get("Result", r.get("result", "PASS"))).upper()
        if "PASS" in status_val:
            status_clean = "PASS"
        else:
            status_clean = "FAIL"
            
        tc_name = r.get("testCase", r.get("title", r.get("Test Case", "")))
        tc_id = r.get("testCaseId", r.get("id", r.get("TC#", "")))
        if tc_id and not tc_name.startswith(tc_id):
            tc_name = f"{tc_id} {tc_name}"
            
        category = r.get("category", r.get("Category", ""))
        
        # Duration handling
        duration_val = r.get("duration", r.get("Duration (s)", r.get("measuredLatency", "0.0")))
        # Strip " ms" or " s" if present
        if isinstance(duration_val, str):
            if " ms" in duration_val:
                try:
                    duration_val = round(float(duration_val.replace(" ms", "").strip()) / 1000.0, 3)
                except ValueError:
                    pass
            elif " s" in duration_val:
                duration_val = duration_val.replace(" s", "").strip()
            else:
                try:
                    duration_val = float(duration_val)
                except ValueError:
                    pass
                    
        details = r.get("details", r.get("Details", r.get("scenario", "Passed successfully")))
        timestamp = r.get("timestamp", r.get("Timestamp", r.get("Started At", gen_time)))
        if isinstance(timestamp, str) and "T" in timestamp:
            timestamp = timestamp.split(".")[0].replace("T", " ")
            
        ws.cell(row=row_num, column=1, value=idx).alignment = center_align
        ws.cell(row=row_num, column=2, value=tc_name).alignment = left_align
        ws.cell(row=row_num, column=3, value=category).alignment = left_align
        
        status_cell = ws.cell(row=row_num, column=4, value=status_clean)
        status_cell.alignment = center_align
        if status_clean == "PASS":
            status_cell.font = pass_font
            status_cell.fill = pass_fill
        else:
            status_cell.font = fail_font
            status_cell.fill = fail_fill
            
        ws.cell(row=row_num, column=5, value=duration_val).alignment = center_align
        ws.cell(row=row_num, column=6, value=details).alignment = left_align
        ws.cell(row=row_num, column=7, value=timestamp).alignment = center_align
        
        # Apply borders and default font
        for col_idx in range(1, 8):
            c = ws.cell(row=row_num, column=col_idx)
            if col_idx != 4: # Status already has font
                c.font = normal_font
            c.border = thin_border
            
    # Set row heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[4].height = 24
    for r_num in range(5, 5 + len(records)):
        ws.row_dimensions[r_num].height = 20
        
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 22
    
    os.makedirs(os.path.dirname(os.path.abspath(excel_path)), exist_ok=True)
    wb.save(excel_path)
    print(f"Generated beautifully styled report: {excel_path}")

if __name__ == "__main__":
    main()
